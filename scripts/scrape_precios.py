#!/usr/bin/env python3
"""
Scraper de Índice de Precios Productor (IPP) e Índice de Precios al Por Mayor (IPM)
Fuente: Instituto Nacional de Estadística (INE) — Bolivia
Base 2016=100

Descarga los Excel del INE y genera JSON estandarizado para el dashboard y embeds.
"""

import json, os, re, sys, time
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# ── Configuración ──────────────────────────────────────────────────────────────

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DOWNLOAD_DIR = Path(__file__).resolve().parent / "downloads"

# Páginas del INE desde donde se RESUELVEN dinámicamente los enlaces de descarga.
# El INE rota los share links (host nube/nimbus + ID) con cada publicación, por eso
# se leen de la página en cada corrida (ver resolver_fuentes). Los enlaces de abajo
# son solo el FALLBACK si el parseo de la página falla. Última verif.: 2026-06-23
PAGINA_IPP = "https://www.ine.gob.bo/index.php/cuadros-estadisticos-ipp/"
PAGINA_IPM = "https://www.ine.gob.bo/index.php/cuadros-estadisticos-ipm/"

FUENTES_IPP = {
    "ipp_general": "https://nube.ine.gob.bo/index.php/s/jiPDzh0nsOiDGY0/download",
    "ipp_grupos":  "https://nube.ine.gob.bo/index.php/s/RbTQhRDB6bpuPWx/download",
}
FUENTES_IPM = {
    "ipm_general":      "https://nube.ine.gob.bo/index.php/s/woA93UUgzbGqzDC/download",
    "ipm_nacional":     "https://nube.ine.gob.bo/index.php/s/WkIa9YYkgWogcJr/download",
    "ipm_importado":    "https://nube.ine.gob.bo/index.php/s/DOBUlEcw9yiLYgt/download",
    "ipm_origen_grupo": "https://nube.ine.gob.bo/index.php/s/xLG383UFEOlt8Lv/download",
}

MESES = {
    "Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
    "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,"Noviembre":11,"Diciembre":12
}

IPP_PONDERACIONES = {
    "Agricolas": 10.27,
    "Industria Manufacturera": 35.33,
    "Otros Minerales y Gas Natural": 12.49,
    "Pecuaria": 4.78,
    "Pesca": 0.11,
    "Servicios": 37.02,
}

# ── Resolución dinámica de enlaces del INE ──────────────────────────────────────

def _sin_acentos(t: str) -> str:
    t = t.lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        t = t.replace(a, b)
    return t


def _extraer_links(html: str) -> list:
    """Devuelve [(texto_normalizado, url), ...] de los anchors a nube/nimbus.ine.gob.bo."""
    links = []
    patron = r'<a[^>]+href="([^"]*(?:nube|nimbus)\.ine\.gob\.bo[^"]*)"[^>]*>(.*?)</a>'
    for m in re.finditer(patron, html, re.I | re.S):
        url = m.group(1).replace("&amp;", "&")
        texto = re.sub(r"<[^>]+>", " ", m.group(2))
        texto = _sin_acentos(re.sub(r"\s+", " ", texto).strip())
        links.append((texto, url))
    return links


def _match(texto: str, key: str) -> bool:
    """Mapea el texto de un enlace del INE a una clave de fuente (IPP/IPM)."""
    t = texto
    reglas = {
        # IPP — página de cuadros del Índice de Precios Productor
        "ipp_grupos":  "general" in t and "grupo" in t,
        "ipp_general": "general" in t and "grupo" not in t,
        # IPM — página de cuadros del Índice de Precios al por Mayor
        "ipm_origen_grupo": "general" in t and "origen" in t and "grupo" in t,
        "ipm_nacional":     "general" in t and "origen" in t and "nacional" in t,
        "ipm_importado":    "general" in t and "origen" in t and "importado" in t,
        "ipm_general":      "general" in t and "origen" not in t and "grupo" not in t,
    }
    return reglas.get(key, False)


def _fetch_html(url: str) -> str:
    import urllib.request, ssl
    ctx = ssl.create_default_context()
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, context=ctx, timeout=60) as resp:
        return resp.read().decode("utf-8", errors="replace")


def resolver_fuentes() -> dict:
    """
    Lee las páginas de cuadros del INE y resuelve los enlaces de descarga actuales
    por el texto de cada enlace. Cae al enlace hardcodeado (fallback) si la página
    no se puede leer o falta alguna clave. Devuelve {clave: url} para las 6 fuentes.
    """
    def resolver(pagina: str, fallback: dict) -> dict:
        try:
            links = _extraer_links(_fetch_html(pagina))
        except Exception as e:
            print(f"  [WARN] No se pudo leer {pagina}: {e} — usando fallback", file=sys.stderr)
            return dict(fallback)
        resuelto = {}
        for texto, url in links:
            for key in fallback:
                if key not in resuelto and _match(texto, key):
                    resuelto[key] = url
                    break
        for key, url in fallback.items():
            if key not in resuelto:
                print(f"  [WARN] '{key}' no encontrado en la página — usando fallback", file=sys.stderr)
                resuelto[key] = url
            else:
                origen = "INE" if resuelto[key] != fallback[key] else "INE (=fallback)"
                print(f"    · {key}: {origen}")
        return resuelto

    sources = {}
    sources.update(resolver(PAGINA_IPP, FUENTES_IPP))
    sources.update(resolver(PAGINA_IPM, FUENTES_IPM))
    return sources


# ── Descarga ───────────────────────────────────────────────────────────────────

def download_all(sources: dict):
    """Descarga todos los Excel del INE con reintentos."""
    import urllib.request, ssl
    ctx = ssl.create_default_context()
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    max_retries = 3
    retry_delays = [10, 30, 60]

    for name, url in sources.items():
        dest = DOWNLOAD_DIR / f"{name}.xlsx"
        print(f"  Descargando {name}...")
        for attempt in range(max_retries):
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, context=ctx, timeout=90) as resp:
                    dest.write_bytes(resp.read())
                print(f"    ✓ {dest.stat().st_size / 1024:.0f} KB")
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    delay = retry_delays[attempt]
                    print(f"    ✗ Intento {attempt + 1}/{max_retries}: {e} — reintentando en {delay}s...")
                    time.sleep(delay)
                else:
                    print(f"    ✗ Error tras {max_retries} intentos: {e}")
                    sys.exit(1)


# ── Parsers ────────────────────────────────────────────────────────────────────

def safe_float(v):
    """Convierte valor a float, None si no es numérico."""
    if v is None:
        return None
    try:
        f = float(v)
        return round(f, 6) if abs(f) < 1e10 else None
    except (ValueError, TypeError):
        return None


def parse_general(filepath, prefix):
    """
    Parsea ipp_general.xlsx o ipm_general.xlsx.
    Estructura: filas=meses, columnas=años, hojas=índice/var_mensual/var_acumulada/var_12_meses
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_map = {}
    for sn in wb.sheetnames:
        sl = sn.upper()
        if "NDICE MENSUAL" in sl or "NDICE MEN" in sl:
            sheet_map["indice"] = sn
        elif "VAR MENSUAL" in sl or "VAR MEN" in sl:
            sheet_map["var_mensual"] = sn
        elif "VAR ACUMULADA" in sl or "VAR ACU" in sl:
            sheet_map["var_acumulada"] = sn
        elif "VAR 12" in sl:
            sheet_map["var_12_meses"] = sn

    series = {}
    for key, sheet_name in sheet_map.items():
        ws = wb[sheet_name]
        years = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(5, c).value
            if v and str(v).strip().isdigit():
                years.append((c, int(str(v).strip())))

        for col, year in years:
            for r in range(7, 19):
                mes_name = ws.cell(r, 1).value
                if not mes_name or str(mes_name).strip() not in MESES:
                    continue
                mes = MESES[str(mes_name).strip()]
                fecha = f"{year}-{mes:02d}"
                val = safe_float(ws.cell(r, col).value)

                if fecha not in series:
                    series[fecha] = {"fecha": fecha}
                series[fecha][key] = val

    result = sorted(series.values(), key=lambda x: x["fecha"])
    result = [r for r in result if any(v is not None for k, v in r.items() if k != "fecha")]
    return result


def parse_grupos(filepath):
    """
    Parsea ipp_grupos.xlsx — datos por grandes grupos de actividad.
    Estructura: filas=grupos, columnas=año×mes (113 cols para 2017-2026)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_keys = {
        "indice": None,
        "var_mensual": None,
        "var_acumulada": None,
        "var_12_meses": None,
    }
    for sn in wb.sheetnames:
        sl = sn.upper()
        if "BOL INDICE" in sl or ("NDICE" in sl and "VAR" not in sl):
            sheet_keys["indice"] = sn
        elif "VAR MENSUAL" in sl:
            sheet_keys["var_mensual"] = sn
        elif "VAR ACUMULADA" in sl:
            sheet_keys["var_acumulada"] = sn
        elif "VAR 12" in sl:
            sheet_keys["var_12_meses"] = sn

    # Parse column headers: year in row 5, month in row 6
    ws_ref = wb[sheet_keys["indice"]]
    col_dates = {}
    current_year = None
    for c in range(3, ws_ref.max_column + 1):
        yr = ws_ref.cell(5, c).value
        if yr and str(yr).strip().isdigit():
            current_year = int(str(yr).strip())
        mn = ws_ref.cell(6, c).value
        if mn and current_year:
            mn_str = str(mn).strip().upper()
            mes_map = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
                       "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
            if mn_str in mes_map:
                col_dates[c] = f"{current_year}-{mes_map[mn_str]:02d}"

    # Parse group names from row 8 onwards
    groups = {}
    for r in range(8, ws_ref.max_row + 1):
        name = ws_ref.cell(r, 2).value
        if not name:
            name = ws_ref.cell(r, 1).value
        if not name or "Fuente" in str(name):
            break
        name = str(name).strip()
        if "GENERAL" in name.upper():
            continue
        groups[r] = name

    # Build series per group
    result = {}
    for key, sheet_name in sheet_keys.items():
        if not sheet_name:
            continue
        ws = wb[sheet_name]
        for row_idx, group_name in groups.items():
            if group_name not in result:
                result[group_name] = {}
            for col_idx, fecha in col_dates.items():
                val = safe_float(ws.cell(row_idx, col_idx).value)
                if fecha not in result[group_name]:
                    result[group_name][fecha] = {"fecha": fecha}
                result[group_name][fecha][key] = val

    output = []
    for group_name, dates_dict in result.items():
        series = sorted(dates_dict.values(), key=lambda x: x["fecha"])
        series = [s for s in series if any(v is not None for k, v in s.items() if k != "fecha")]
        output.append({
            "grupo": group_name,
            "ponderacion": IPP_PONDERACIONES.get(group_name, 0),
            "series": series
        })

    output.sort(key=lambda x: -x["ponderacion"])
    return output


def parse_ipm_origen(filepath):
    """
    Parsea ipm_origen_grupo.xlsx — General, Nacional (Agrícola, Manufactura), Importado (Agrícola, Manufactura)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_keys = {}
    for sn in wb.sheetnames:
        sl = sn.upper()
        if "BOL INDICE" in sl or ("NDICE" in sl and "VAR" not in sl):
            sheet_keys["indice"] = sn
        elif "VAR MENSUAL" in sl:
            sheet_keys["var_mensual"] = sn
        elif "VAR ACUMULADA" in sl:
            sheet_keys["var_acumulada"] = sn
        elif "VAR 12" in sl:
            sheet_keys["var_12_meses"] = sn

    ws_ref = wb[sheet_keys["indice"]]

    # Parse column dates
    col_dates = {}
    current_year = None
    for c in range(2, ws_ref.max_column + 1):
        yr = ws_ref.cell(5, c).value
        if yr and str(yr).strip().isdigit():
            current_year = int(str(yr).strip())
        mn = ws_ref.cell(6, c).value
        if mn and current_year:
            mn_str = str(mn).strip().upper()
            mes_map = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
                       "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
            if mn_str in mes_map:
                col_dates[c] = f"{current_year}-{mes_map[mn_str]:02d}"

    # Parse category names, prefixing subgroups with their origin
    categories = {}
    current_origin = ""
    for r in range(8, ws_ref.max_row + 1):
        name = ws_ref.cell(r, 1).value
        if not name or "Fuente" in str(name):
            break
        name = str(name).strip()
        if name.startswith("ORIGEN"):
            current_origin = name
            categories[r] = name
        elif name == "GENERAL":
            categories[r] = name
        else:
            categories[r] = f"{current_origin} — {name}" if current_origin else name

    # Build series
    result = {}
    for key, sheet_name in sheet_keys.items():
        if not sheet_name:
            continue
        ws = wb[sheet_name]
        for row_idx, cat_name in categories.items():
            if cat_name not in result:
                result[cat_name] = {}
            for col_idx, fecha in col_dates.items():
                val = safe_float(ws.cell(row_idx, col_idx).value)
                if fecha not in result[cat_name]:
                    result[cat_name][fecha] = {"fecha": fecha}
                result[cat_name][fecha][key] = val

    output = []
    for cat_name, dates_dict in result.items():
        series = sorted(dates_dict.values(), key=lambda x: x["fecha"])
        series = [s for s in series if any(v is not None for k, v in s.items() if k != "fecha")]
        output.append({"categoria": cat_name, "series": series})

    return output


# ── Main ───────────────────────────────────────────────────────────────────────

def save_json(data, filename):
    """Guarda JSON con formato compacto."""
    dest = DATA_DIR / filename
    with open(dest, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    size = dest.stat().st_size / 1024
    print(f"  ✓ {filename} ({size:.1f} KB, {len(data) if isinstance(data, list) else 'obj'} items)")


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print("═══ Scraper IPP/IPM — INE Bolivia ═══\n")

    # 1. Resolver enlaces actuales del INE (rotan con cada publicación) y descargar
    print("1. Resolviendo enlaces del INE...")
    sources = resolver_fuentes()
    print("   Descargando Excel del INE...")
    download_all(sources)

    # 2. Procesar IPP
    print("\n2. Procesando IPP...")
    ipp_general = parse_general(DOWNLOAD_DIR / "ipp_general.xlsx", "ipp")
    print(f"  IPP general: {len(ipp_general)} meses ({ipp_general[0]['fecha']} → {ipp_general[-1]['fecha']})")

    ipp_grupos = parse_grupos(DOWNLOAD_DIR / "ipp_grupos.xlsx")
    print(f"  IPP grupos: {len(ipp_grupos)} grupos")
    for g in ipp_grupos:
        print(f"    - {g['grupo']} ({g['ponderacion']}%): {len(g['series'])} meses")

    # 3. Procesar IPM
    print("\n3. Procesando IPM...")
    ipm_general = parse_general(DOWNLOAD_DIR / "ipm_general.xlsx", "ipm")
    print(f"  IPM general: {len(ipm_general)} meses ({ipm_general[0]['fecha']} → {ipm_general[-1]['fecha']})")

    ipm_origen = parse_ipm_origen(DOWNLOAD_DIR / "ipm_origen_grupo.xlsx")
    print(f"  IPM origen/grupo: {len(ipm_origen)} categorías")
    for cat in ipm_origen:
        print(f"    - {cat['categoria']}: {len(cat['series'])} meses")

    # 4. Generar JSON combinado IPP vs IPM (var 12 meses para comparación)
    print("\n4. Generando comparativo...")
    comparativo = {}
    for entry in ipp_general:
        f = entry["fecha"]
        comparativo[f] = {"fecha": f, "ipp_indice": entry.get("indice"), "ipp_var12": entry.get("var_12_meses")}
    for entry in ipm_general:
        f = entry["fecha"]
        if f not in comparativo:
            comparativo[f] = {"fecha": f}
        comparativo[f]["ipm_indice"] = entry.get("indice")
        comparativo[f]["ipm_var12"] = entry.get("var_12_meses")

    comparativo_list = sorted(comparativo.values(), key=lambda x: x["fecha"])

    # 5. Metadata
    now = datetime.now(timezone.utc).isoformat()
    last_ipp = ipp_general[-1]["fecha"] if ipp_general else None
    last_ipm = ipm_general[-1]["fecha"] if ipm_general else None

    metadata = {
        "actualizado": now,
        "fuente": "Instituto Nacional de Estadística (INE) — Bolivia",
        "base": "2016=100",
        "urls": {
            "ipp_cuadros": "https://www.ine.gob.bo/index.php/cuadros-estadisticos-ipp/",
            "ipm_cuadros": "https://www.ine.gob.bo/index.php/cuadros-estadisticos-ipm/",
            "ipp_metodologia": "https://www.ine.gob.bo/index.php/informacion-tecnica-ipp/",
            "ipm_metodologia": "https://www.ine.gob.bo/index.php/informacion-tecnica-ipm/",
        },
        "ultimo_dato_ipp": last_ipp,
        "ultimo_dato_ipm": last_ipm,
        "ipp": {
            "nombre": "Índice de Precios Productor",
            "definicion": "Mide la variación promedio de precios de bienes y servicios a la salida del lugar de producción.",
            "formula": "Laspeyres con ponderaciones fijas",
            "productos": 223,
            "informantes": 9613,
            "cobertura": "9 departamentos, ciudades capitales, conurbaciones y UPAs rurales",
            "grupos": {g["grupo"]: g["ponderacion"] for g in ipp_grupos}
        },
        "ipm": {
            "nombre": "Índice de Precios al Por Mayor",
            "definicion": "Mide la variación de precios en el canal mayorista, diferenciando bienes nacionales e importados.",
            "formula": "Laspeyres con ponderaciones fijas",
            "origenes": ["Nacional", "Importado"],
        },
        "frecuencia": "Mensual",
    }

    # 6. Guardar
    print("\n5. Guardando JSON...")
    save_json(ipp_general, "ipp_general.json")
    save_json(ipp_grupos, "ipp_grupos.json")
    save_json(ipm_general, "ipm_general.json")
    save_json(ipm_origen, "ipm_origen.json")
    save_json(comparativo_list, "comparativo.json")
    save_json(metadata, "metadata.json")

    # Resumen
    last = ipp_general[-1] if ipp_general else {}
    last_ipm_d = ipm_general[-1] if ipm_general else {}
    print(f"\n═══ Resumen ═══")
    print(f"  IPP último: {last.get('fecha')} — índice {last.get('indice', 0):.1f}, var12m {last.get('var_12_meses', 0):.1f}%")
    print(f"  IPM último: {last_ipm_d.get('fecha')} — índice {last_ipm_d.get('indice', 0):.1f}, var12m {last_ipm_d.get('var_12_meses', 0):.1f}%")
    print(f"  Archivos: {len(list(DATA_DIR.glob('*.json')))} JSON en data/")


if __name__ == "__main__":
    main()
